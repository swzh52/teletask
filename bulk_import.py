"""
bulk_import.py
批量导入「关键词自动回复」与「定时推送任务」
支持 .xlsx / .csv / .json 三种格式

设计原则：导入字段与网页手动输入的字段完全一致，确保手动添加和批量导入能力对等。
"""
import csv, io, logging
from datetime import datetime
import database as db

log = logging.getLogger(__name__)

# ============================================================
# 列名规范（与手动输入表单字段完全对齐）
# ============================================================
#
# ─── 关键词表（一个 pattern 可以写多行，合并为多条回复）───
#   pattern              关键词文本 / 正则
#   match                匹配方式：contains / exact / regex（默认 contains）
#   mode                 触发模式：random / all（默认 random）
#
#   reply_type           回复类型：text / photo / video / audio / document /
#                        animation / voice / sticker
#   reply_text           文本内容（支持 HTML 富文本）
#   reply_file_id        Telegram file_id（非 text 类型必填）
#   reply_caption        图片/视频等的说明文字
#
#   delete_after_seconds 发送后自动删除延迟（秒）。0 或留空 = 不自动删除
#   delete_after_value   便捷写法：数值（与 _unit 搭配；若已填 _seconds 则忽略）
#   delete_after_unit    便捷写法：单位秒数（1=秒 60=分 3600=时 86400=天）
#
#   expire_after_seconds 关键词自动到期时长（秒），到期后停用
#   expire_after_value   便捷写法：数值
#   expire_after_unit    便捷写法：单位秒数
#
#   start_at             开始生效时间（"YYYY-MM-DD HH:MM:SS" 或 "YYYY-MM-DD HH:MM"）
#   chat_ids             监控哪些群组；留空 = 所有群组
#                        多个用逗号/空格/分号分隔；JSON 里可用数组
#   active               1=启用（默认） 0=停用
#
# ─── 定时任务表 ───
#   name                 任务名称
#   chat_id              目标 Chat ID
#   cron                 周期任务的 cron 表达式（5 字段，如 "0 9 * * *"）
#                        一次性任务填具体时间（"YYYY-MM-DD HH:MM:SS"）
#   once                 0=周期执行（默认） 1=一次性
#   msg_type             消息类型：text / photo / video / audio / document /
#                        animation / voice / sticker
#   msg_text             文本内容（支持 HTML 富文本）
#   msg_file_id          Telegram file_id（非 text 类型必填）
#   msg_caption          图片/视频等的说明文字
#   delete_after_seconds 发送后自动删除延迟（秒）
#   delete_after_value   便捷写法：数值
#   delete_after_unit    便捷写法：单位秒数
#   active               1=启用（默认） 0=停用

KW_COLS = [
    "pattern", "match", "mode",
    "reply_type", "reply_text", "reply_file_id", "reply_caption",
    "delete_after_seconds", "delete_after_value", "delete_after_unit",
    "expire_after_seconds", "expire_after_value", "expire_after_unit",
    "start_at", "chat_ids", "active",
]

SC_COLS = [
    "name", "chat_id", "cron", "once",
    "msg_type", "msg_text", "msg_file_id", "msg_caption",
    "delete_after_seconds", "delete_after_value", "delete_after_unit",
    "active",
]


# ============================================================
# 解析辅助
# ============================================================
def _safe_int(v, default=None):
    if v is None or v == "":
        return default
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return default


def _safe_str(v, default=""):
    if v is None:
        return default
    s = str(v).strip()
    return s if s else default


def _parse_chat_ids(v):
    """解析 '-1001,-1002' 或 JSON 数组 为 int 列表"""
    if v is None or v == "":
        return []
    if isinstance(v, list):
        ids = []
        for p in v:
            try:
                ids.append(int(float(str(p).strip())))
            except (ValueError, TypeError):
                pass
        return ids
    s = str(v).replace("，", ",").replace(";", ",").replace(" ", ",")
    ids = []
    for p in s.split(","):
        p = p.strip()
        if not p:
            continue
        try:
            ids.append(int(float(p)))
        except ValueError:
            pass
    return ids


def _resolve_seconds(row, seconds_key, value_key, unit_key):
    """
    统一处理「总秒数」和「数值+单位」两种写法。优先级：
      1. 直接指定 *_seconds
      2. *_value + *_unit（单位秒数）合成
      3. 都没有 → None
    """
    sec = _safe_int(row.get(seconds_key))
    if sec is not None:
        return sec if sec > 0 else None
    val = _safe_int(row.get(value_key))
    if val is None or val <= 0:
        return None
    unit = _safe_int(row.get(unit_key), 1) or 1
    return val * unit


def _normalize_start_at(s):
    """接受 "YYYY-MM-DD HH:MM" 或 "YYYY-MM-DD HH:MM:SS"；非法返回 None"""
    s = _safe_str(s)
    if not s:
        return None
    s = s.replace("T", " ")         # 兼容 datetime-local 原样粘贴
    if len(s) == 16:
        s += ":00"
    try:
        datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
        return s
    except ValueError:
        return None


# ============================================================
# 读取文件
# ============================================================
def _read_rows(filename, content_bytes):
    """返回 list[dict]；列名统一 lower-case strip"""
    ext = filename.lower().rsplit(".", 1)[-1]
    if ext == "csv":
        text = None
        for enc in ("utf-8-sig", "utf-8", "gbk"):
            try:
                text = content_bytes.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise ValueError("无法识别 CSV 编码，请保存为 UTF-8")
        reader = csv.DictReader(io.StringIO(text))
        rows = []
        for r in reader:
            rows.append({(k or "").strip().lower(): v for k, v in r.items()})
        return rows
    elif ext in ("xlsx", "xls"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError("未安装 openpyxl，请执行: pip install openpyxl")
        wb = load_workbook(io.BytesIO(content_bytes), data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [(str(h).strip().lower() if h is not None else "")
                   for h in next(rows_iter, [])]
        rows = []
        for r in rows_iter:
            if all(v is None or str(v).strip() == "" for v in r):
                continue
            rows.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})
        return rows
    elif ext == "json":
        import json
        try:
            text = content_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content_bytes.decode("utf-8", errors="replace")
        data = json.loads(text)
        if isinstance(data, list):
            items = data
        elif isinstance(data, dict):
            items = data.get("keywords") or data.get("schedules") or data.get("data") or []
        else:
            raise ValueError("JSON 顶层必须是数组或对象")
        rows = []
        for item in items:
            if not isinstance(item, dict):
                continue
            norm = {}
            for k, v in item.items():
                key = (k or "").strip().lower()
                # chat_ids 数组原样透传给 _parse_chat_ids
                norm[key] = v
            rows.append(norm)
        return rows
    else:
        raise ValueError(f"不支持的文件格式: .{ext}，仅支持 .xlsx / .csv / .json")


# ============================================================
# 关键词导入
# ============================================================
def import_keywords(filename, content_bytes):
    """
    返回 (ok_count, fail_count, errors: list[str])
    相同 pattern 的多行会合并为多条回复（按文件中出现的顺序）
    """
    rows = _read_rows(filename, content_bytes)
    if not rows:
        return 0, 0, ["文件为空"]

    groups = {}
    order  = []
    for i, r in enumerate(rows, start=2):
        pattern = _safe_str(r.get("pattern"))
        if not pattern:
            continue
        key = (pattern, _safe_str(r.get("match"), "contains"))
        if key not in groups:
            groups[key] = []
            order.append(key)
        r["_row_num"] = i
        groups[key].append(r)

    ok, fail, errors = 0, 0, []
    for key in order:
        pattern, match = key
        rs = groups[key]
        first = rs[0]
        try:
            # 校验 match 合法值
            if match not in ("contains", "exact", "regex"):
                fail += 1
                errors.append(f"第{first['_row_num']}行 [{pattern}]: match 必须是 contains/exact/regex")
                continue

            replies = []
            for r in rs:
                rtype = _safe_str(r.get("reply_type"), "text")
                replies.append({
                    "reply_type":    rtype,
                    "reply_text":    _safe_str(r.get("reply_text")) or None,
                    "reply_file_id": _safe_str(r.get("reply_file_id")) or None,
                    "reply_caption": _safe_str(r.get("reply_caption")) or None,
                })
            replies = [r for r in replies
                       if r["reply_text"] or r["reply_file_id"] or r["reply_type"] == "sticker"]
            if not replies:
                fail += 1
                errors.append(f"第{first['_row_num']}行 [{pattern}]: 无有效回复内容")
                continue

            mode = _safe_str(first.get("mode"), "random")
            if mode not in ("random", "all"):
                mode = "random"

            das = _resolve_seconds(first, "delete_after_seconds",
                                   "delete_after_value", "delete_after_unit")
            eas = _resolve_seconds(first, "expire_after_seconds",
                                   "expire_after_value", "expire_after_unit")
            start_at = _normalize_start_at(first.get("start_at"))
            chat_ids = _parse_chat_ids(first.get("chat_ids"))
            active   = _safe_int(first.get("active"), 1)

            kid = db.add_keyword(pattern, match, mode, replies,
                                 delete_after_seconds=das,
                                 expire_after_seconds=eas,
                                 start_at=start_at,
                                 chat_ids=chat_ids)
            if active == 0 and kid:
                db.toggle_keyword(kid)
            ok += 1
        except Exception as e:
            fail += 1
            errors.append(f"第{first['_row_num']}行 [{pattern}]: {e}")
    return ok, fail, errors


# ============================================================
# 定时任务导入
# ============================================================
def import_schedules(filename, content_bytes):
    rows = _read_rows(filename, content_bytes)
    if not rows:
        return 0, 0, ["文件为空"]

    valid_types = {"text", "photo", "video", "audio", "document",
                   "animation", "voice", "sticker"}

    ok, fail, errors = 0, 0, []
    for i, r in enumerate(rows, start=2):
        name    = _safe_str(r.get("name"))
        chat_id = _safe_str(r.get("chat_id"))
        cron    = _safe_str(r.get("cron"))
        if not name or not chat_id or not cron:
            if any(r.get(c) for c in SC_COLS):
                fail += 1
                errors.append(f"第{i}行: name/chat_id/cron 缺失")
            continue
        try:
            once = _safe_int(r.get("once"), 0) == 1

            if once:
                cron = cron.replace("T", " ")
                if len(cron) == 16:
                    cron += ":00"
                try:
                    datetime.strptime(cron, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    fail += 1
                    errors.append(f"第{i}行 [{name}]: 一次性任务时间格式错误，应为 YYYY-MM-DD HH:MM:SS")
                    continue
            else:
                if len(cron.split()) != 5:
                    fail += 1
                    errors.append(f"第{i}行 [{name}]: cron 表达式必须是 5 个字段")
                    continue

            msg_type = _safe_str(r.get("msg_type"), "text")
            if msg_type not in valid_types:
                fail += 1
                errors.append(f"第{i}行 [{name}]: msg_type 无效（允许 {'/'.join(sorted(valid_types))}）")
                continue

            das = _resolve_seconds(r, "delete_after_seconds",
                                   "delete_after_value", "delete_after_unit")
            active = _safe_int(r.get("active"), 1)

            sid = db.add_schedule(
                name=name,
                chat_id=chat_id,
                cron=cron,
                msg_type=msg_type,
                msg_text=_safe_str(r.get("msg_text")) or None,
                msg_file_id=_safe_str(r.get("msg_file_id")) or None,
                msg_caption=_safe_str(r.get("msg_caption")) or None,
                once=1 if once else 0,
                delete_after_seconds=das,
                start_at=None,
            )
            # active=0 需立刻停用刚新增的条目，直接用返回的 ID
            if active == 0 and sid:
                db.toggle_schedule(sid)
            ok += 1
        except Exception as e:
            fail += 1
            errors.append(f"第{i}行 [{name}]: {e}")
    return ok, fail, errors


# ============================================================
# 模板导出（CSV）
# ============================================================
def build_keyword_template_csv():
    """关键词 CSV 模板；字段顺序与 KW_COLS 对齐"""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(KW_COLS)
    # 示例 1：同一 pattern 两条文本回复（随机）+ 10 秒自动删除
    w.writerow(["价格", "contains", "random",
                "text", "我们的价格请咨询客服 <b>@support</b>", "", "",
                "10", "", "",
                "", "", "",
                "", "", "1"])
    w.writerow(["价格", "contains", "random",
                "text", "价目表已更新，请查看置顶。", "", "",
                "10", "", "",
                "", "", "",
                "", "", "1"])
    # 示例 2：精确匹配；便捷写法 7 天后到期；限定单个群组
    w.writerow(["早安", "exact", "random",
                "text", "早上好 🌞", "", "",
                "", "", "",
                "", "7", "86400",
                "", "-1001234567890", "1"])
    # 示例 3：图片回复（all 模式），指定生效时间 + 两个群组
    w.writerow(["菜单", "contains", "all",
                "photo", "", "AgACAgIAAxkBAAExxxxxxxxx", "今日菜单，欢迎点单",
                "", "", "",
                "", "", "",
                "2025-06-01 09:00:00", "-1001111111111,-1002222222222", "1"])
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def build_schedule_template_csv():
    """定时任务 CSV 模板；字段顺序与 SC_COLS 对齐"""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(SC_COLS)
    # 示例 1：每天早 9 点推送
    w.writerow(["早报推送", "-1001234567890", "0 9 * * *", "0",
                "text", "<b>早安！</b>新的一天开始啦", "", "",
                "0", "", "", "1"])
    # 示例 2：一次性任务；便捷写法 30 分钟后自删
    w.writerow(["新年祝福", "-1001234567890", "2026-01-01 00:00:00", "1",
                "text", "🎉 新年快乐！", "", "",
                "", "30", "60", "1"])
    # 示例 3：图片定时推送（每周一 10 点）
    w.writerow(["每周菜单", "-1001234567890", "0 10 * * 1", "0",
                "photo", "", "AgACAgIAAxkBAAExxxxxxxxx", "本周菜单",
                "0", "", "", "1"])
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


# ============================================================
# 模板导出（JSON）—— 与 CSV 字段完全对齐
# ============================================================
def build_keyword_template_json():
    import json
    data = {
        "keywords": [
            {
                "pattern": "价格",
                "match": "contains",
                "mode": "random",
                "reply_type": "text",
                "reply_text": "我们的价格请咨询客服 <b>@support</b>",
                "reply_file_id": "",
                "reply_caption": "",
                "delete_after_seconds": 10,
                "delete_after_value": "",
                "delete_after_unit": "",
                "expire_after_seconds": "",
                "expire_after_value": "",
                "expire_after_unit": "",
                "start_at": "",
                "chat_ids": [],
                "active": 1
            },
            {
                "pattern": "价格",
                "match": "contains",
                "mode": "random",
                "reply_type": "text",
                "reply_text": "价目表已更新，请查看置顶。",
                "reply_file_id": "",
                "reply_caption": "",
                "delete_after_seconds": 10,
                "delete_after_value": "",
                "delete_after_unit": "",
                "expire_after_seconds": "",
                "expire_after_value": "",
                "expire_after_unit": "",
                "start_at": "",
                "chat_ids": [],
                "active": 1
            },
            {
                "pattern": "早安",
                "match": "exact",
                "mode": "random",
                "reply_type": "text",
                "reply_text": "早上好 🌞",
                "reply_file_id": "",
                "reply_caption": "",
                "delete_after_seconds": "",
                "delete_after_value": "",
                "delete_after_unit": "",
                "expire_after_seconds": "",
                "expire_after_value": 7,
                "expire_after_unit": 86400,
                "start_at": "",
                "chat_ids": [-1001234567890],
                "active": 1
            },
            {
                "pattern": "菜单",
                "match": "contains",
                "mode": "all",
                "reply_type": "photo",
                "reply_text": "",
                "reply_file_id": "AgACAgIAAxkBAAExxxxxxxxx",
                "reply_caption": "今日菜单，欢迎点单",
                "delete_after_seconds": "",
                "delete_after_value": "",
                "delete_after_unit": "",
                "expire_after_seconds": "",
                "expire_after_value": "",
                "expire_after_unit": "",
                "start_at": "2025-06-01 09:00:00",
                "chat_ids": [-1001111111111, -1002222222222],
                "active": 1
            }
        ]
    }
    return json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")


def build_schedule_template_json():
    import json
    data = {
        "schedules": [
            {
                "name": "早报推送",
                "chat_id": "-1001234567890",
                "cron": "0 9 * * *",
                "once": 0,
                "msg_type": "text",
                "msg_text": "<b>早安！</b>新的一天开始啦",
                "msg_file_id": "",
                "msg_caption": "",
                "delete_after_seconds": 0,
                "delete_after_value": "",
                "delete_after_unit": "",
                "active": 1
            },
            {
                "name": "新年祝福",
                "chat_id": "-1001234567890",
                "cron": "2026-01-01 00:00:00",
                "once": 1,
                "msg_type": "text",
                "msg_text": "🎉 新年快乐！",
                "msg_file_id": "",
                "msg_caption": "",
                "delete_after_seconds": "",
                "delete_after_value": 30,
                "delete_after_unit": 60,
                "active": 1
            },
            {
                "name": "每周菜单",
                "chat_id": "-1001234567890",
                "cron": "0 10 * * 1",
                "once": 0,
                "msg_type": "photo",
                "msg_text": "",
                "msg_file_id": "AgACAgIAAxkBAAExxxxxxxxx",
                "msg_caption": "本周菜单",
                "delete_after_seconds": 0,
                "delete_after_value": "",
                "delete_after_unit": "",
                "active": 1
            }
        ]
    }
    return json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
